"""Adapter for the current official OAIC NDB XLSX resource."""

from __future__ import annotations

import re
from datetime import UTC, date, datetime
from io import BytesIO
from typing import Any

import httpx
from openpyxl import load_workbook

from breachgazette.clients.base import (
    AdapterResult,
    BoundedSourceClient,
    SourceClientError,
    source_snapshot,
)
from breachgazette.contracts import SourceAggregateRecord
from breachgazette.contracts.enums import (
    Completeness,
    PublicationLevel,
    ValueOrigin,
    ValueState,
)
from breachgazette.contracts.models import ObservedValue, RecordProvenance
from breachgazette.utils import normalize_text, sha256_hex

DATASET_ID = "5781dc17-2ad0-4dd1-bced-01c544943ce3"
PACKAGE_URL = f"https://data.gov.au/data/api/3/action/package_show?id={DATASET_ID}"
DATASET_URL = f"https://data.gov.au/data/dataset/{DATASET_ID}"
EXPECTED_SHEETS = (
    "Cover page",
    "NDB by month",
    "Individuals affected",
    "Personal information",
    "Source of breach",
    "Top 5 sectors by source",
    "Time to identify by sector",
    "Time to identify by source",
    "Time to notify by sector",
    "Time to notify by source",
)
MONTHS = {
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
}
ROOT_CAUSES = {
    "Human error",
    "Malicious or criminal attack",
    "System fault",
    "Other",
    "Currently unknown",
}


def _resource_period(name: str) -> tuple[date, date]:
    match = re.search(
        r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})\s+to\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        name,
    )
    if not match:
        raise SourceClientError("OAIC resource name omitted the reporting period")
    return (
        datetime.strptime(match.group(1), "%d %B %Y").date(),
        datetime.strptime(match.group(2), "%d %b %Y").date(),
    )


def _numeric(value: Any) -> int | float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        text = value.replace(",", "").replace("%", "").strip()
        try:
            number = float(text)
            return int(number) if number.is_integer() else number
        except ValueError:
            return None
    return None


class OaicNdbAdapter:
    source_id = "oaic_ndb"
    adapter_version = "1.0"
    normalization_version = "1.0"
    max_records = 2_000

    def __init__(self, *, transport: httpx.BaseTransport | None = None) -> None:
        self.transport = transport

    def collect(self, *, observed_at: datetime | None = None) -> AdapterResult:
        observed_at = observed_at or datetime.now(UTC)
        with BoundedSourceClient(
            allowed_origins={"https://data.gov.au"},
            allowed_path_prefixes=("/data/api/3/action/", "/data/dataset/"),
            max_response_bytes=10_000_000,
            transport=self.transport,
        ) as client:
            package, _headers, _url = client.get_json(PACKAGE_URL)
            if not isinstance(package, dict) or package.get("success") is not True:
                raise SourceClientError("Data.gov.au package response was not successful")
            dataset = package.get("result")
            if not isinstance(dataset, dict):
                raise SourceClientError("Data.gov.au package omitted its result")
            if dataset.get("license_id") != "cc-by-4.0":
                raise SourceClientError("OAIC source licence changed")
            resources = dataset.get("resources")
            if not isinstance(resources, list):
                raise SourceClientError("OAIC dataset resources are not a list")
            xlsx_resources = [
                resource
                for resource in resources
                if isinstance(resource, dict)
                and str(resource.get("format", "")).upper() == "XLSX"
                and str(resource.get("name", "")).startswith("NDB Data ")
            ]
            if len(xlsx_resources) != 1:
                raise SourceClientError("OAIC dataset did not expose exactly one reviewed XLSX")
            resource = xlsx_resources[0]
            resource_url = str(resource.get("url", ""))
            content, headers, final_url = client.get_bytes(
                resource_url,
                accept=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            )
            content_type = headers.get("Content-Type", "").lower()
            if (
                "spreadsheetml" not in content_type
                and "octet-stream" not in content_type
                and not final_url.endswith(".xlsx")
            ):
                raise SourceClientError("OAIC resource returned an unexpected content type")

        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if tuple(workbook.sheetnames) != EXPECTED_SHEETS:
            raise SourceClientError("OAIC workbook sheet schema changed")
        period_start, period_end = _resource_period(str(resource["name"]))
        checksum = sha256_hex(content)
        revision = f"{resource.get('id')}:{dataset.get('metadata_modified')}"
        records: list[RecordProvenance] = []
        source_notes = [
            "Data are based on information self-reported by notifying entities.",
            "Statistics may be revised.",
            "Notification counts are not automatically unique real-world incidents.",
        ]

        for worksheet in workbook.worksheets[1:]:
            current_parent: str | None = None
            population_scope = "Notifications received under the NDB scheme"
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                label_raw = row[0] if row else None
                value_raw = row[1] if len(row) > 1 else None
                if not isinstance(label_raw, str):
                    continue
                label = normalize_text(label_raw, maximum=500)
                if "world-wide" in label.lower() or "worldwide" in label.lower():
                    population_scope = "Individuals worldwide affected, as source-reported"
                    current_parent = label
                    continue
                if "australian" in label.lower() and "affected" in label.lower():
                    population_scope = "Australians affected, as source-reported"
                    current_parent = label
                    continue
                value = _numeric(value_raw)
                if value is None:
                    if value_raw in {None, ""} and row_number > 3:
                        current_parent = label
                    continue
                if label in MONTHS:
                    current_parent = label
                    dimension = "notifications_by_month"
                    parent = None
                elif worksheet.title == "NDB by month":
                    dimension = "notifications_by_month_and_source"
                    parent = current_parent
                elif worksheet.title == "Source of breach" and label in ROOT_CAUSES:
                    current_parent = label
                    dimension = "source_of_breach"
                    parent = None
                else:
                    dimension = re.sub(r"[^a-z0-9]+", "_", worksheet.title.casefold()).strip("_")
                    parent = current_parent
                unit = "percent" if isinstance(value_raw, str) and "%" in value_raw else "count"
                rounding = ValueState.ESTIMATED if unit == "percent" else ValueState.PRESENT
                source_record_id = (
                    f"{worksheet.title.strip()}:{row_number}:"
                    f"{sha256_hex([label, parent, value])[:12]}"
                )
                records.append(
                    SourceAggregateRecord(
                        source_id=self.source_id,
                        source_record_id=source_record_id,
                        source_url=DATASET_URL,
                        source_revision=revision,
                        source_checksum=checksum,
                        source_completeness=Completeness.COMPLETE,
                        source_retrieval_time=observed_at,
                        local_first_observed_time=observed_at,
                        local_last_observed_time=observed_at,
                        parser_version=self.adapter_version,
                        normalization_version=self.normalization_version,
                        limitations=["Aggregate metrics do not identify notifying organizations."],
                        regulator="Office of the Australian Information Commissioner",
                        reporting_scheme="Notifiable Data Breaches scheme",
                        publication_level=PublicationLevel.NATIONAL_AGGREGATE,
                        reporting_period_start=period_start,
                        reporting_period_end=period_end,
                        dimension=dimension,
                        category=label,
                        parent_category=parent,
                        value=ObservedValue(
                            value=value,
                            origin=ValueOrigin.SOURCE_OBSERVED,
                            state=rounding,
                            source_label=str(value_raw),
                        ),
                        unit=unit,
                        population_scope=population_scope,
                        rounding_state=rounding,
                        source_notes=source_notes,
                    )
                )
        workbook.close()
        if not records or len(records) > self.max_records:
            raise SourceClientError("OAIC metric count was empty or exceeded its bound")
        snapshot = source_snapshot(
            source_id=self.source_id,
            retrieved_at=observed_at,
            revision=revision,
            checksum=checksum,
            completeness=Completeness.COMPLETE,
            discovered=len(records),
            accepted=len(records),
            rejected=0,
            bounded_limit=self.max_records,
            source_updated_at=datetime.fromisoformat(
                str(dataset["metadata_modified"]).replace("Z", "+00:00")
            ),
            notes=source_notes,
        )
        return AdapterResult(source_id=self.source_id, records=records, snapshot=snapshot)
