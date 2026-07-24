"""Bounded, privacy-minimised adapter for UK ICO incident-trend data."""

from __future__ import annotations

import calendar
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from email.utils import parsedate_to_datetime
from io import BytesIO
from pathlib import PurePosixPath
from urllib.parse import urljoin, urlparse
from zipfile import BadZipFile, ZipFile

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from breachgazette.clients.base import (
    AdapterResult,
    BoundedSourceClient,
    SourceClientError,
    source_snapshot,
)
from breachgazette.contracts import SourceAggregateRecord
from breachgazette.contracts.enums import (
    Completeness,
    PublicationLevel,
    ValueOrigin,
    ValueState,
)
from breachgazette.contracts.models import ObservedValue, RecordProvenance
from breachgazette.utils import normalize_text, sha256_hex

SOURCE_URL = (
    "https://ico.org.uk/action-weve-taken/complaints-and-concerns-data-sets/"
    "data-security-incident-trends/"
)
WORKBOOK_PATH = re.compile(
    r"^/media2/[a-z0-9]+/"
    r"data-security-incidents-trends-q1-2019-to-q([1-4])-(\d{4})\.xlsx$"
)
EXPECTED_HEADERS = (
    "BI Reference",
    "Year",
    "Quarter",
    "Data Subject Type",
    "Data Type",
    "Decision Taken",
    "Incident Category",
    "Incident Type",
    "No. Data Subjects Affected",
    "Sector",
    "Time Taken to Report",
)
EARLIEST_COMPARABLE_PERIOD = (2019, 2)
REQUIRED_PAGE_NOTES = (
    "the data starts at q2 2019",
    "some reports hold multiple characteristics",
    "open government licence v3.0",
)
VOCABULARIES = {
    "data_subject_type": {
        "Children",
        "Customers or prospective customers",
        "Employees",
        "Patients",
        "Students",
        "Subscribers",
        "Unknown",
        "Users",
        "Vulnerable adults",
    },
    "data_type": {
        "Basic personal identifiers",
        "Criminal convictions or offences",
        "Data revealing racial or ethnic origin",
        "Economic and financial data",
        "Gender Reassignment Data",
        "Genetic or biometric data",
        "Health data",
        "Identification data",
        "Location data",
        "Official documents",
        "Political opinions",
        "Religious or philosophical beliefs",
        "Sex life data",
        "Sexual orientation data",
        "Trade union membership",
        "Unknown",
    },
    "decision_taken": {
        "Informal Action Taken",
        "Investigation Pursued",
        "No Further Action",
        "Not Yet Assigned",
        "Regulatory action taken",
    },
    "incident_category": {"Cyber", "Non Cyber"},
    "incident_type": {
        "Alteration of personal data",
        "Brute Force",
        "Cryptographic flaw",
        "Data emailed to incorrect recipient",
        "Data of wrong data subject shown in client portal",
        "Data posted or faxed to incorrect recipient",
        "Denial of service",
        "Failure to redact",
        "Failure to use bcc",
        "Hardware/software misconfiguration",
        "Incorrect disposal of hardware",
        "Incorrect disposal of paperwork",
        "Loss/theft of device containing personal data",
        "Loss/theft of paperwork or data left in insecure location",
        "Malware",
        "Not Provided",
        "Other cyber incident",
        "Other non-cyber incident",
        "Phishing",
        "Ransomware",
        "Unauthorised access",
        "Verbal disclosure of personal data",
    },
    "affected_population_band": {
        "1 to 9",
        "10 to 99",
        "100 to 1k",
        "1k to 10k",
        "10k to 100k",
        "100k and above",
        "Unknown",
    },
    "sector": {
        "Central Government",
        "Charitable and voluntary",
        "Education and childcare",
        "Finance, insurance and credit",
        "General business",
        "Health",
        "Justice",
        "Land or property services",
        "Legal",
        "Local government",
        "Marketing",
        "Media",
        "Membership association",
        "Online Technology and Telecoms",
        "Political",
        "Regulators",
        "Religious",
        "Retail and manufacture",
        "Social care",
        "Transport and leisure",
        "Unassigned",
        "Unknown",
        "Utilities",
    },
    "time_to_report": {
        "Less than 24 hours",
        "24 hours to 72 hours",
        "72 hours to 1 week",
        "More than 1 week",
    },
}


@dataclass(frozen=True, slots=True)
class IcoResource:
    url: str
    filename: str
    coverage_end: date


@dataclass(slots=True)
class IcoReport:
    periods: set[tuple[int, int]] = field(default_factory=set)
    categories: dict[str, set[str]] = field(
        default_factory=lambda: {dimension: set() for dimension in VOCABULARIES}
    )


def _quarter_end(year: int, quarter: int) -> date:
    month = quarter * 3
    return date(year, month, calendar.monthrange(year, month)[1])


def _quarter_start(year: int, quarter: int) -> date:
    return date(year, ((quarter - 1) * 3) + 1, 1)


def _periods_through(end_year: int, end_quarter: int) -> set[tuple[int, int]]:
    periods: set[tuple[int, int]] = set()
    year, quarter = 2019, 1
    while (year, quarter) <= (end_year, end_quarter):
        periods.add((year, quarter))
        if quarter == 4:
            year, quarter = year + 1, 1
        else:
            quarter += 1
    return periods


def _select_workbook(html: str) -> IcoResource:
    soup = BeautifulSoup(html, "html.parser")
    page_text = " ".join(soup.stripped_strings).casefold()
    if any(note not in page_text for note in REQUIRED_PAGE_NOTES):
        raise SourceClientError("ICO source semantics or licence notice changed")
    candidates: dict[str, IcoResource] = {}
    for anchor in soup.select("[href], [x-href]"):
        href = anchor.get("href") or anchor.get("x-href")
        if not isinstance(href, str):
            continue
        url = urljoin(SOURCE_URL, href)
        parsed = urlparse(url)
        match = WORKBOOK_PATH.fullmatch(parsed.path)
        if (
            parsed.scheme != "https"
            or parsed.netloc != "ico.org.uk"
            or parsed.query
            or parsed.fragment
            or match is None
        ):
            continue
        end_quarter, end_year = int(match.group(1)), int(match.group(2))
        if (end_year, end_quarter) < (2025, 4):
            raise SourceClientError("ICO workbook coverage moved behind the reviewed boundary")
        candidates[url] = IcoResource(
            url=url,
            filename=PurePosixPath(parsed.path).name,
            coverage_end=_quarter_end(end_year, end_quarter),
        )
    if len(candidates) != 1:
        raise SourceClientError("ICO page did not expose exactly one reviewed workbook")
    return next(iter(candidates.values()))


def _source_updated_at(headers: httpx.Headers) -> datetime:
    value = headers.get("Last-Modified")
    if not value:
        raise SourceClientError("ICO workbook omitted Last-Modified")
    try:
        parsed = parsedate_to_datetime(value)
    except (TypeError, ValueError) as exc:
        raise SourceClientError("ICO workbook Last-Modified changed") from exc
    if parsed.tzinfo is None:
        raise SourceClientError("ICO workbook Last-Modified omitted its timezone")
    return parsed.astimezone(UTC)


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if not entries or len(entries) > 100:
                raise SourceClientError("ICO workbook archive exceeded its file-count bound")
            total_size = 0
            for entry in entries:
                path = PurePosixPath(entry.filename)
                if (
                    path.is_absolute()
                    or ".." in path.parts
                    or entry.flag_bits & 0x1
                    or entry.file_size > 100_000_000
                ):
                    raise SourceClientError("ICO workbook archive contained an unsafe entry")
                total_size += entry.file_size
            if total_size > 120_000_000:
                raise SourceClientError("ICO workbook archive exceeded its expanded-size bound")
    except BadZipFile as exc:
        raise SourceClientError("ICO workbook was not a valid XLSX archive") from exc


def _cell_text(value: object, *, field_name: str) -> str:
    if value is None:
        raise SourceClientError(f"ICO {field_name} field was blank")
    try:
        return normalize_text(str(value), maximum=200)
    except ValueError as exc:
        raise SourceClientError(f"ICO {field_name} field exceeded its bound") from exc


def _parse_workbook(
    content: bytes,
    *,
    coverage_end: date,
    max_rows: int,
    max_reports: int,
) -> tuple[dict[str, IcoReport], int, int, list[str]]:
    _validate_xlsx_archive(content)
    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if workbook.sheetnames != ["in"]:
            raise SourceClientError("ICO workbook sheet schema changed")
        worksheet = workbook["in"]
        if (
            worksheet.max_column != len(EXPECTED_HEADERS)
            or worksheet.max_row is None
            or worksheet.max_row < 2
            or worksheet.max_row > max_rows + 1
        ):
            raise SourceClientError("ICO workbook dimensions exceeded their reviewed bounds")
        rows = worksheet.iter_rows(values_only=True)
        raw_headers = next(rows, None)
        if raw_headers is None:
            raise SourceClientError("ICO workbook omitted its header row")
        headers = tuple(_cell_text(value, field_name="header") for value in raw_headers)
        if headers != EXPECTED_HEADERS:
            raise SourceClientError("ICO workbook schema changed")

        reports: dict[str, IcoReport] = {}
        raw_rows = 0
        for row_number, raw_row in enumerate(rows, start=2):
            raw_rows += 1
            if raw_rows > max_rows or len(raw_row) != len(EXPECTED_HEADERS):
                raise SourceClientError("ICO workbook row count or width exceeded its bound")
            reference = _cell_text(raw_row[0], field_name="reference")
            if re.fullmatch(r"BI[1-9][0-9]*", reference) is None:
                raise SourceClientError(f"ICO workbook row {row_number} reference changed")
            year_value = raw_row[1]
            if isinstance(year_value, bool) or not isinstance(year_value, int):
                raise SourceClientError("ICO workbook year format changed")
            quarter_label = _cell_text(raw_row[2], field_name="quarter")
            quarter_match = re.fullmatch(r"Qtr ([1-4])", quarter_label)
            if quarter_match is None:
                raise SourceClientError("ICO workbook quarter format changed")
            quarter = int(quarter_match.group(1))
            if (year_value, quarter) < (2019, 1) or _quarter_end(
                year_value, quarter
            ) > coverage_end:
                raise SourceClientError("ICO workbook row fell outside its coverage boundary")

            report = reports.setdefault(reference, IcoReport())
            if len(reports) > max_reports:
                raise SourceClientError("ICO workbook exceeded its unique-reference bound")
            report.periods.add((year_value, quarter))
            for index, dimension in enumerate(VOCABULARIES, start=3):
                category = _cell_text(raw_row[index], field_name=dimension)
                if category not in VOCABULARIES[dimension]:
                    raise SourceClientError(f"ICO {dimension} vocabulary changed")
                report.categories[dimension].add(category)

        if raw_rows == 0 or not reports:
            raise SourceClientError("ICO workbook contained no incident reports")
        end_period = (coverage_end.year, (coverage_end.month - 1) // 3 + 1)
        observed_periods = {period for report in reports.values() for period in report.periods}
        if observed_periods != _periods_through(*end_period):
            raise SourceClientError("ICO workbook quarter coverage changed")

        pre_window = {
            reference for reference, report in reports.items() if report.periods == {(2019, 1)}
        }
        scoped = {
            reference: report
            for reference, report in reports.items()
            if reference not in pre_window
        }
        conflicts = sorted(
            reference for reference, report in scoped.items() if len(report.periods) != 1
        )
        accepted = {
            reference: report for reference, report in scoped.items() if reference not in conflicts
        }
        return accepted, raw_rows, len(pre_window), conflicts
    finally:
        workbook.close()


def _aggregate_records(
    reports: dict[str, IcoReport],
    *,
    resource: IcoResource,
    checksum: str,
    revision: str,
    observed_at: datetime,
) -> list[RecordProvenance]:
    first_period = EARLIEST_COMPARABLE_PERIOD
    last_period = (resource.coverage_end.year, (resource.coverage_end.month - 1) // 3 + 1)
    overall_start = _quarter_start(*first_period)
    overall_end = _quarter_end(*last_period)
    report_count = len(reports)
    common_notes = [
        "Counts use unique source report references rather than workbook rows.",
        "Q1 2019 is excluded because the source identifies Q2 2019 as its comparable start.",
        "A source reference spanning conflicting quarters is excluded rather than resolved.",
        "Category membership totals can exceed unique reports because reports can have "
        "multiple characteristics.",
        "Raw workbook rows and report-level characteristic combinations are not retained.",
    ]
    records: list[RecordProvenance] = []

    def add_metric(
        *,
        dimension: str,
        category: str,
        value: int,
        unit: str,
        period_start: date = overall_start,
        period_end: date = overall_end,
        denominator: int | None = None,
    ) -> None:
        identity = sha256_hex([dimension, category])[:20]
        records.append(
            SourceAggregateRecord(
                source_id="united_kingdom_ico",
                source_record_id=f"uk:ico:{dimension}:{identity}",
                source_url=SOURCE_URL,
                source_revision=revision,
                source_checksum=checksum,
                source_completeness=Completeness.PARTIAL,
                source_retrieval_time=observed_at,
                local_first_observed_time=observed_at,
                local_last_observed_time=observed_at,
                parser_version="1.0",
                normalization_version="1.0",
                limitations=common_notes,
                regulator="Information Commissioner's Office",
                reporting_scheme="UK GDPR personal data breach reports to the ICO",
                publication_level=PublicationLevel.ANONYMIZED_NOTIFICATION,
                reporting_period_start=period_start,
                reporting_period_end=period_end,
                dimension=dimension,
                category=category,
                value=ObservedValue(
                    value=value,
                    origin=ValueOrigin.CALCULATED,
                    state=ValueState.PRESENT,
                ),
                unit=unit,
                population_scope=(
                    "In-scope unique report references in the published ICO workbook"
                ),
                denominator=denominator,
                rounding_state=ValueState.PRESENT,
                source_notes=common_notes,
            )
        )

    add_metric(
        dimension="unique_reports",
        category="All unambiguous in-scope report references",
        value=report_count,
        unit="unique_source_report_references",
    )
    quarter_counts = Counter(next(iter(report.periods)) for report in reports.values())
    for (year, quarter), count in sorted(quarter_counts.items()):
        add_metric(
            dimension="reporting_quarter",
            category=f"{year} Q{quarter}",
            value=count,
            unit="unique_source_report_references",
            period_start=_quarter_start(year, quarter),
            period_end=_quarter_end(year, quarter),
        )
    for dimension in VOCABULARIES:
        counts = Counter(
            category for report in reports.values() for category in report.categories[dimension]
        )
        for category, count in sorted(counts.items()):
            add_metric(
                dimension=dimension,
                category=category,
                value=count,
                unit="report_category_memberships",
                denominator=report_count,
            )
    return records


class UnitedKingdomIcoAdapter:
    source_id = "united_kingdom_ico"
    adapter_version = "1.0"
    normalization_version = "1.0"
    minimum_source_reports = 70_000
    max_reports = 100_000
    max_rows = 250_000
    max_response_bytes = 20_000_000

    def __init__(self, *, transport: httpx.BaseTransport | None = None) -> None:
        self.transport = transport

    def collect(self, *, observed_at: datetime | None = None) -> AdapterResult:
        observed_at = observed_at or datetime.now(UTC)
        with BoundedSourceClient(
            allowed_origins={"https://ico.org.uk"},
            allowed_path_prefixes=(
                "/action-weve-taken/complaints-and-concerns-data-sets/",
                "/media2/",
            ),
            max_response_bytes=self.max_response_bytes,
            total_deadline_seconds=180,
            transport=self.transport,
            user_agent=(
                "Mozilla/5.0 (compatible; BreachGazette/0.1; +https://github.com/slicedearth)"
            ),
        ) as client:
            html, _page_headers, _page_url = client.get_text(SOURCE_URL)
            resource = _select_workbook(html)
            content, headers, final_url = client.get_bytes(
                resource.url,
                accept=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            )
        content_type = headers.get("Content-Type", "").casefold()
        if (
            "spreadsheetml" not in content_type
            and "octet-stream" not in content_type
            and not final_url.endswith(".xlsx")
        ):
            raise SourceClientError("ICO workbook returned an unexpected content type")
        source_updated_at = _source_updated_at(headers)
        reports, raw_rows, excluded_before_window, conflicts = _parse_workbook(
            content,
            coverage_end=resource.coverage_end,
            max_rows=self.max_rows,
            max_reports=self.max_reports,
        )
        if len(reports) < self.minimum_source_reports:
            raise SourceClientError("ICO unique-report count fell below its reviewed bound")
        checksum = sha256_hex(content)
        revision = f"ico-workbook:{resource.filename}:{source_updated_at.isoformat()}"
        records = _aggregate_records(
            reports,
            resource=resource,
            checksum=checksum,
            revision=revision,
            observed_at=observed_at,
        )
        conflict_label = "reference" if len(conflicts) == 1 else "references"
        notes = [
            f"Validated {raw_rows} bounded workbook rows transiently.",
            f"Excluded {excluded_before_window} Q1 2019 references outside the source's "
            "stated comparable window.",
            f"Accepted {len(reports)} unambiguous in-scope report references.",
            f"Excluded {len(conflicts)} {conflict_label} with conflicting reporting periods.",
            f"Published {len(records)} privacy-minimised aggregate cells.",
            "Raw workbook rows and report-level characteristic combinations were not retained.",
        ]
        snapshot = source_snapshot(
            source_id=self.source_id,
            retrieved_at=observed_at,
            revision=revision,
            checksum=checksum,
            completeness=Completeness.PARTIAL,
            discovered=len(reports),
            accepted=len(reports),
            rejected=0,
            bounded_limit=self.max_reports,
            source_updated_at=source_updated_at,
            notes=notes,
        )
        return AdapterResult(
            source_id=self.source_id,
            records=records,
            snapshot=snapshot,
        )
